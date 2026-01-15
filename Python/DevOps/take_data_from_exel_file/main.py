import openpyxl  # Library for working with Excel (.xlsx) files

# python3 -m venv .tmp
# source .tmp/bin/activate
# pip install openpyxl

# Load the Excel file
inv_file = openpyxl.load_workbook("inventory.xlsx")
# Select the worksheet that contains product data
product_list = inv_file["Sheet1"]

# Dictionary to store the number of products per supplier
# key   -> supplier name
# value -> product count
product_per_supplier = {}

# Start from row 2 to skip the header row
# Excel rows are 1-based (indexing starts at 1)
for row_index in range(2, product_list.max_row + 1):
    # Read supplier name from column 4
    supplier_name = product_list.cell(row_index, 4).value

    # calculation number of products per supplier
    if supplier_name in product_per_supplier:
        product_per_supplier[supplier_name] += 1
    else:
        product_per_supplier[supplier_name] = 1

print(product_per_supplier)

# calculation total value of inventory per supplier
total_value_per_supplier = {}
for row_index in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(row_index, 4).value
    inventory = product_list.cell(row_index, 2).value
    price = product_list.cell(row_index, 3).value
    inventory_price = inventory * price

    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = (
            current_total_value + inventory_price
        )
    else:
        total_value_per_supplier[supplier_name] = inventory_price

    # Write the inventory value of this product into column 5
    inventory_price_cell = product_list.cell(row_index, 5)
    inventory_price_cell.value = inventory_price

print(total_value_per_supplier)

inv_file.save("inventory_new.xlsx")
inv_file.close()
