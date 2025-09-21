import xlsxwriter

# python3 -m venv .tmp
# source .tmp/bin/activate
# pip install xlsxwriter

# Create a workbook and add a worksheet.
workbook = xlsxwriter.Workbook("tmp.xlsx")
worksheet = workbook.add_worksheet()

# Some data we want to write to the worksheet.
expenses = (
    ["Rent", 1000],
    ["Gas", 100],
    ["Food", 300],
    ["Gym", 50],
)

# Start from the first cell. Rows and columns are zero indexed.
row = 0
col = 0

# Iterate over the data and write it out row by row.
for item, cost in expenses:
    worksheet.write(row, col, item)
    worksheet.write(row, col + 1, cost)
    row += 1

# Write a total using a formula.
worksheet.write(row, 0, "Total")
worksheet.write(row, 1, "=SUM(B1:B4)")

workbook.close()

print("The content of the file is:")

import openpyxl

wb = openpyxl.load_workbook("tmp.xlsx")
ws = wb.active

for row in ws.iter_rows(values_only=True):
    print(row)